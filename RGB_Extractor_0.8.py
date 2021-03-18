import os
from tkinter import *
import tkinter.filedialog
import imageio
from pathlib import Path
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def get_R(p):
    
    pic = imageio.imread(p)
    
    length = pic.shape[0]
    width = pic.shape[1]
    R = 0

    for i in range(length):
        for j in range(width):
            R = R + pic[ i, j, 0]
            
    return R

def get_G(p):
    
    pic = imageio.imread(p)
    
    length = pic.shape[0]
    width = pic.shape[1]
    G = 0

    for i in range(length):
        for j in range(width):
            G = G + pic[ i, j, 1]
    return G

def get_B(p):
    
    pic = imageio.imread(p)
    
    length = pic.shape[0]
    width = pic.shape[1]
    B = 0
    
    for i in range(length):
        for j in range(width):
            B = B + pic[ i, j, 2]
            
    return B

def get_date_taken(p):
    return Image.open(p)._getexif()[36867]

def alg(name,path):
    filename = name + ".xlsx"
    wb = Workbook()
    dataSheet = wb.create_sheet('Data',0)
    ref = wb['Sheet']
    wb.remove(ref)
    active = wb['Data']
    
    pathlist = Path(path).glob('**/*.JPG')

    current = dataSheet.cell(row= 1 , column = 1) 
    current.value = "date-time"
    current = dataSheet.cell(row= 1 , column = 2) 
    current.value = "Total R"
    current = dataSheet.cell(row= 1 , column = 3) 
    current.value = "Total G"
    current = dataSheet.cell(row= 1 , column = 4) 
    current.value = "Total B"
    
    line = 2
    
    for path in pathlist:
        path_in_str = str(path)
        current = dataSheet.cell(row= line , column = 1) 
        current.value = get_date_taken(path)
        current = dataSheet.cell(row= line , column = 2) 
        current.value = get_R(path_in_str)
        current = dataSheet.cell(row= line , column = 3) 
        current.value = get_G(path_in_str)
        current = dataSheet.cell(row= line , column = 4) 
        current.value = get_B(path_in_str)

        wb.save(filename)
        line = line + 1

window = Tk()
window.title("RGB Data into Excel 0.4")
window.geometry('350x200+200+150')

lbl = Label(window, text="Hello")
lbl.grid(column=0, row=0)

txt = Entry(window,width=10)
txt.grid(column=1, row=0)

def run():
    filename = txt.get()
    curr_directory = os.getcwd()
    path = tkinter.filedialog.askdirectory()
    alg(filename,path)
    
def clicked():
    res = "Welcome to " + txt.get()
    lbl.configure(text= res)

btn = Button(window, text="Enter", command=clicked)
btn.grid(column=2, row=0)

btn2 = Button(window, text="Choose Directory and Run", command=run)
btn2.grid(columnspan = 2)

window.mainloop()
